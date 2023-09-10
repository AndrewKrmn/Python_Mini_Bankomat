import telebot
from telebot import types
import openpyxl
wBook = openpyxl.load_workbook('Bank_history.xlsx')
sheet = wBook.active
next_row = sheet.max_row + 1
TOKEN = "6659756402:AAGwYhea35qBozf0nED94U3UVko0BARAyy4"
bot = telebot.TeleBot(TOKEN)
description_add_cash = "Список поповнень на карту"
description_remove_cash = "Список зняття коштів з карти"
cash = "0"

count_func = 0


@bot.message_handler(commands=["start"])
def start_message(message):
    markup = types.ReplyKeyboardMarkup(row_width=2)
    item1 = types.KeyboardButton("Переглянути вміст коштів на картці")
    item2 = types.KeyboardButton("Поповнити картку")
    item3 = types.KeyboardButton("Зняти кошти з картки")
    item4 = types.KeyboardButton("Взяти кредит")
    item5 = types.KeyboardButton("Вивести історію Транзакцій")
    markup.add(item1, item2, item3, item4,item5)
    bot.send_message(
        message.chat.id, "Bank Bot\nВибери операцію :", reply_markup=markup
    )


@bot.message_handler(
    func=lambda message: message.text == "Переглянути вміст коштів на картці"
)
def show_wallet(message):
    global cash
    bot.send_message(message.chat.id, f"Твій баланс :\n{cash}")


@bot.message_handler(func=lambda message: message.text == "Поповнити картку")
def add_wallet(message):
    bot.send_message(message.chat.id, "Впиши суму коштів для поповнення :")
    bot.register_next_step_handler(message, add_cash)


def add_cash(message):
    global next_row
    global wBook
    global cash
    count_func += 1
    amount = float(cash) + float(message.text)
    cash = amount
    next_row = sheet.max_row + 1
    sheet['A'+ str(next_row)] = description_add_cash
    sheet['C'+ str(next_row)] = f"Сума поповненнь : {message.text}"
    sheet['E'+ str(next_row)] = f"На балансі : {cash}"
    wBook.save('Bank_history.xlsx')
    bot.send_message(message.chat.id, f"Поповнено коштів на суму :{message.text}")


@bot.message_handler(func=lambda message: message.text == "Зняти кошти з картки")
def remove_wallet(message):
    bot.send_message(message.chat.id, "Впиши суму коштів для зняття :")
    bot.register_next_step_handler(message, remove_cash)


def remove_cash(message):
    global next_row
    global wBook
    global sheet
    global description_remove_cash
    global count_func
    global cash
    count_func += 1
    amount = float(cash) - float(message.text)
    cash = amount
    next_row = sheet.max_row + 1
    sheet['A'+ str(next_row)] = description_remove_cash
    sheet['C'+ str(next_row)] = f"Сума зняття : {message.text}"
    sheet['E'+ str(next_row)] = f"На балансі : {cash}"
    wBook.save('Bank_history.xlsx')
    bot.send_message(message.chat.id, f"Знято коштів на суму :{message.text}")


@bot.message_handler(func=lambda message: message.text == "Взяти кредит")
def get_credit(message):
    global count_func
    global cash
    if count_func >= 5 and cash > 50000:
        bot.send_message(
            message.chat.id,
            f"Ваш кредитний ліміт опираючись на {count_func} транзакцію(-ії),становить 100000 ",
        )
    elif count_func >= 5 and cash > 10000:
        bot.send_message(
            message.chat.id,
            f"Ваш кредитний ліміт опираючись на {count_func} транзакцію(-ії),становить 15000 ",
        )
    elif count_func >= 5 and cash > 1000:
        bot.send_message(
            message.chat.id,
            f"Ваш кредитний ліміт опираючись на {count_func} транзакцію(-ії),становить 5000 ",
        )
    else:
        bot.send_message(message.chat.id, "Кредит недоступний!")

@bot.message_handler(
    func=lambda message: message.text == "Вивести історію Транзакцій")
def show_list(message):
    global wBook
    global sheet
    show = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = ', '.join(map(str, row)) 
        show.append(row_data)

    data_text = '\n\n'.join(show)
    bot.send_message(message.chat.id, f"Данные из Excel:\n\n{data_text}")

bot.infinity_polling()
